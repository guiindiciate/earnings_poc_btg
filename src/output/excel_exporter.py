"""Excel exporter — generates multi-tab Excel workbooks from the database.

Four tabs are produced:

1. **KPIs_Core**         — core financial metrics across reporting periods.
2. **KPIs_Operacionais** — free-form operational KPIs.
3. **Comparativo**       — cross-ticker comparison for a single period.
4. **Metadados**         — extraction audit trail.
"""

from __future__ import annotations

import logging
import os
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from config.settings import OUTPUT_PATH
from src.output.excel_template import (
    FMT_BRL,
    FMT_DATE,
    FMT_MULTIPLE,
    FMT_PCT,
    HEADER_FILL,
    HEADER_FONT,
    THIN_BORDER,
    auto_fit_columns,
    style_data_cell,
    style_header_cell,
    style_variation_cell,
)
from src.storage.database import init_db, session_scope
from src.storage.models import EarningsRecord, ExtractionLog, OperationalKPI

logger = logging.getLogger(__name__)

# ── Metric display configuration ──────────────────────────────────────────────

# (display_label, record_attribute, number_format, is_variation)
_CORE_METRIC_ROWS: list[tuple[str, str, str, bool]] = [
    ("Receita Líquida (R$ MM)", "receita_liquida", FMT_BRL, False),
    ("  Var. a/a (%)", "receita_liquida_var_aa", FMT_PCT, True),
    ("Lucro Bruto (R$ MM)", "lucro_bruto", FMT_BRL, False),
    ("  Margem Bruta (%)", "lucro_bruto_margem", FMT_PCT, False),
    ("  Var. a/a (%)", "lucro_bruto_var_aa", FMT_PCT, True),
    ("EBITDA (R$ MM)", "ebitda", FMT_BRL, False),
    ("  Margem EBITDA (%)", "ebitda_margem", FMT_PCT, False),
    ("  Var. a/a (%)", "ebitda_var_aa", FMT_PCT, True),
    ("EBIT (R$ MM)", "ebit", FMT_BRL, False),
    ("  Margem EBIT (%)", "ebit_margem", FMT_PCT, False),
    ("  Var. a/a (%)", "ebit_var_aa", FMT_PCT, True),
    ("Lucro Líquido (R$ MM)", "lucro_liquido", FMT_BRL, False),
    ("  Margem Líquida (%)", "lucro_liquido_margem", FMT_PCT, False),
    ("  Var. a/a (%)", "lucro_liquido_var_aa", FMT_PCT, True),
    ("Lucro Líq. Controlador (R$ MM)", "lucro_liquido_controlador", FMT_BRL, False),
    ("ROE (%)", "roe", FMT_PCT, False),
    ("  Var. a/a (%)", "roe_var_aa", FMT_PCT, True),
    ("ROIC (%)", "roic", FMT_PCT, False),
    ("ROA (%)", "roa", FMT_PCT, False),
    ("Margem EBITDA (%)", "margem_ebitda", FMT_PCT, False),
    ("Margem Líquida (%)", "margem_liquida", FMT_PCT, False),
    ("Ativo Total (R$ MM)", "ativo_total", FMT_BRL, False),
    ("Patrimônio Líquido (R$ MM)", "patrimonio_liquido", FMT_BRL, False),
    ("Dívida Bruta (R$ MM)", "divida_bruta", FMT_BRL, False),
    ("Caixa e Equivalentes (R$ MM)", "caixa_equivalentes", FMT_BRL, False),
    ("Dívida Líquida (R$ MM)", "divida_liquida", FMT_BRL, False),
    ("Alavancagem DL/EBITDA (x)", "alavancagem_dl_ebitda", FMT_MULTIPLE, False),
    ("CFO (R$ MM)", "cfo", FMT_BRL, False),
    ("Capex (R$ MM)", "capex", FMT_BRL, False),
    ("FCL (R$ MM)", "fcl", FMT_BRL, False),
    ("Conversão de Caixa (%)", "conversao_caixa", FMT_PCT, False),
    ("Capital de Giro (R$ MM)", "capital_de_giro", FMT_BRL, False),
    ("PMR (dias)", "pmr", "#,##0", False),
    ("PMP (dias)", "pmp", "#,##0", False),
    ("PMIE (dias)", "pmie", "#,##0", False),
]


def _sort_key(periodo: str) -> tuple[int, int]:
    """Return ``(ano, trimestre)`` for chronological sorting."""
    try:
        q_str, y_str = periodo.upper().split("T")
        q = int(q_str)
        y = int(y_str)
        y = y + 2000 if y < 100 else y
        return y, q
    except (ValueError, AttributeError):
        return 9999, 9


def _write_kpis_core(wb: Workbook, records: list[EarningsRecord]) -> None:
    """Write the KPIs_Core tab."""
    ws = wb.create_sheet("KPIs_Core")

    if not records:
        ws.append(["Sem dados disponíveis"])
        return

    # Header row: "Métrica" + one column per period in chronological order
    periods = sorted({r.periodo for r in records}, key=_sort_key)
    period_to_record: dict[str, EarningsRecord] = {}
    for r in records:
        # Keep the most recent extraction per period
        if r.periodo not in period_to_record or r.created_at > period_to_record[r.periodo].created_at:
            period_to_record[r.periodo] = r

    # Row 1 — header
    header = ["Métrica"] + periods
    ws.append(header)
    for col_idx, _ in enumerate(header, start=1):
        style_header_cell(ws.cell(row=1, column=col_idx))
    ws.freeze_panes = "B2"

    # Data rows
    for row_idx, (label, attr, fmt, is_var) in enumerate(_CORE_METRIC_ROWS, start=2):
        ws.cell(row=row_idx, column=1, value=label)
        label_cell = ws.cell(row=row_idx, column=1)
        label_cell.font = Font(name="Calibri", size=10, bold=not label.startswith("  "))
        label_cell.border = THIN_BORDER

        for col_idx, period in enumerate(periods, start=2):
            record = period_to_record.get(period)
            value = getattr(record, attr, None) if record else None
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if is_var:
                style_variation_cell(cell, value, row_idx)
            else:
                style_data_cell(cell, row_idx, number_format=fmt)

    auto_fit_columns(ws)


def _write_kpis_operacionais(wb: Workbook, ticker: str, records: list[EarningsRecord]) -> None:
    """Write the KPIs_Operacionais tab."""
    ws = wb.create_sheet("KPIs_Operacionais")

    periods = sorted({r.periodo for r in records}, key=_sort_key)

    # Fetch operational KPIs from DB
    init_db()
    with session_scope() as session:
        from sqlalchemy import select

        kpi_rows = list(
            session.scalars(
                select(OperationalKPI)
                .where(OperationalKPI.ticker == ticker)
                .order_by(OperationalKPI.nome_kpi, OperationalKPI.periodo)
            ).all()
        )

    if not kpi_rows:
        ws.append(["Sem KPIs operacionais disponíveis"])
        return

    # Build pivot: {kpi_name: {period: (valor, unidade)}}
    pivot: dict[str, dict[str, tuple[Any, str | None]]] = {}
    for kpi in kpi_rows:
        pivot.setdefault(kpi.nome_kpi, {})[kpi.periodo] = (kpi.valor, kpi.unidade)

    kpi_names = sorted(pivot.keys())

    # Header
    header = ["KPI Operacional", "Unidade"] + periods
    ws.append(header)
    for col_idx, _ in enumerate(header, start=1):
        style_header_cell(ws.cell(row=1, column=col_idx))
    ws.freeze_panes = "C2"

    # Data rows
    for row_idx, kpi_name in enumerate(kpi_names, start=2):
        period_data = pivot[kpi_name]
        # Try to get the unit from any available period
        unit = next((v[1] for v in period_data.values() if v[1]), None)

        ws.cell(row=row_idx, column=1, value=kpi_name).border = THIN_BORDER
        ws.cell(row=row_idx, column=2, value=unit).border = THIN_BORDER

        for col_idx, period in enumerate(periods, start=3):
            val = period_data.get(period, (None, None))[0]
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            style_data_cell(cell, row_idx)

    auto_fit_columns(ws)


def _write_comparativo(
    wb: Workbook, tickers: list[str], periodo: str, records_by_ticker: dict[str, EarningsRecord | None]
) -> None:
    """Write the Comparativo tab."""
    ws = wb.create_sheet("Comparativo")

    header = ["Métrica"] + tickers
    ws.append(header)
    for col_idx, _ in enumerate(header, start=1):
        style_header_cell(ws.cell(row=1, column=col_idx))
    ws.freeze_panes = "B2"

    for row_idx, (label, attr, fmt, is_var) in enumerate(_CORE_METRIC_ROWS, start=2):
        ws.cell(row=row_idx, column=1, value=label).border = THIN_BORDER
        for col_idx, ticker in enumerate(tickers, start=2):
            record = records_by_ticker.get(ticker)
            value = getattr(record, attr, None) if record else None
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if is_var:
                style_variation_cell(cell, value, row_idx)
            else:
                style_data_cell(cell, row_idx, number_format=fmt)

    auto_fit_columns(ws)


def _write_metadados(wb: Workbook, ticker: str) -> None:
    """Write the Metadados (audit) tab."""
    ws = wb.create_sheet("Metadados")

    headers = [
        "Ticker", "Período", "Ano", "Trimestre", "Data Extração",
        "Arquivo Origem", "Confiança Score", "Revisão Manual", "Status",
    ]
    ws.append(headers)
    for col_idx, _ in enumerate(headers, start=1):
        style_header_cell(ws.cell(row=1, column=col_idx))

    init_db()
    with session_scope() as session:
        from sqlalchemy import select

        logs = list(
            session.scalars(
                select(ExtractionLog)
                .where(ExtractionLog.ticker == ticker)
                .order_by(ExtractionLog.created_at.desc())
            ).all()
        )
        records = list(
            session.scalars(
                select(EarningsRecord)
                .where(EarningsRecord.ticker == ticker)
                .order_by(EarningsRecord.ano, EarningsRecord.trimestre)
            ).all()
        )

    # Merge by (ticker, periodo)
    meta_by_periodo: dict[str, dict] = {}
    for r in records:
        meta_by_periodo[r.periodo] = {
            "ticker": r.ticker,
            "periodo": r.periodo,
            "ano": r.ano,
            "trimestre": r.trimestre,
            "data_extracao": r.data_extracao,
            "arquivo_origem": r.arquivo_origem,
            "confianca_score": r.confianca_score,
            "revisao_manual": r.revisao_manual,
            "status": r.status,
        }

    for row_idx, (_, meta) in enumerate(
        sorted(meta_by_periodo.items(), key=lambda x: _sort_key(x[0])), start=2
    ):
        row = [
            meta["ticker"],
            meta["periodo"],
            meta["ano"],
            meta["trimestre"],
            meta["data_extracao"],
            meta["arquivo_origem"],
            meta["confianca_score"],
            bool(meta["revisao_manual"]),
            meta["status"],
        ]
        ws.append(row)
        for col_idx in range(1, len(row) + 1):
            style_data_cell(ws.cell(row=row_idx, column=col_idx), row_idx)

    auto_fit_columns(ws)


# ── Public API ─────────────────────────────────────────────────────────────────


def export_to_excel(ticker: str, output_path: str | None = None) -> str:
    """Generate a historical multi-period Excel workbook for a single ticker.

    Parameters
    ----------
    ticker:
        Stock ticker symbol (case-insensitive).
    output_path:
        Destination file path.  Defaults to ``OUTPUT_PATH/<TICKER>_resultados.xlsx``.

    Returns
    -------
    str
        Absolute path to the generated Excel file.
    """
    ticker = ticker.upper()
    init_db()

    with session_scope() as session:
        from sqlalchemy import select

        records = list(
            session.scalars(
                select(EarningsRecord)
                .where(EarningsRecord.ticker == ticker)
                .order_by(EarningsRecord.ano, EarningsRecord.trimestre)
            ).all()
        )

    if not records:
        logger.warning("No records found for ticker %s", ticker)

    if output_path is None:
        out_dir = Path(OUTPUT_PATH)
        out_dir.mkdir(parents=True, exist_ok=True)
        output_path = str(out_dir / f"{ticker}_resultados.xlsx")

    wb = Workbook()
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    _write_kpis_core(wb, records)
    _write_kpis_operacionais(wb, ticker, records)
    _write_metadados(wb, ticker)

    wb.save(output_path)
    logger.info("Excel exported to %s", output_path)
    return output_path


def export_comparative(
    tickers: list[str], periodo: str, output_path: str | None = None
) -> str:
    """Generate a cross-ticker comparison Excel for a single reporting period.

    Parameters
    ----------
    tickers:
        List of ticker symbols to compare.
    periodo:
        Reporting period label (e.g. ``"4T25"``).
    output_path:
        Destination file path.  Defaults to ``OUTPUT_PATH/comparativo_<PERIODO>.xlsx``.

    Returns
    -------
    str
        Absolute path to the generated Excel file.
    """
    tickers = [t.upper() for t in tickers]
    periodo = periodo.upper()
    init_db()

    records_by_ticker: dict[str, EarningsRecord | None] = {}
    with session_scope() as session:
        from sqlalchemy import select

        for ticker in tickers:
            record = session.scalar(
                select(EarningsRecord)
                .where(
                    EarningsRecord.ticker == ticker,
                    EarningsRecord.periodo == periodo,
                )
                .order_by(EarningsRecord.created_at.desc())
                .limit(1)
            )
            records_by_ticker[ticker] = record

    if output_path is None:
        out_dir = Path(OUTPUT_PATH)
        out_dir.mkdir(parents=True, exist_ok=True)
        output_path = str(out_dir / f"comparativo_{periodo}.xlsx")

    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    _write_comparativo(wb, tickers, periodo, records_by_ticker)

    wb.save(output_path)
    logger.info("Comparative Excel exported to %s", output_path)
    return output_path
