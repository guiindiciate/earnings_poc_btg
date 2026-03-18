"""Excel visual template — colours, fonts, number formats and layout helpers.

All formatting constants and helper functions used by :mod:`~src.output.excel_exporter`
are centralised here to make visual changes easy without touching business logic.
"""

from __future__ import annotations

import openpyxl.cell.cell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# ── Colour palette ─────────────────────────────────────────────────────────────

HEADER_BG = "1F3864"       # dark navy — column / row headers background
HEADER_FG = "FFFFFF"       # white text on header
ROW_ALT_BG = "F2F2F2"      # light grey for alternating rows
VAR_POSITIVE_FG = "00B050"  # green for positive variations
VAR_NEGATIVE_FG = "FF0000"  # red for negative variations
VAR_NEUTRAL_FG = "595959"   # dark grey when value is zero / null

# ── Fills ──────────────────────────────────────────────────────────────────────

HEADER_FILL = PatternFill(fill_type="solid", fgColor=HEADER_BG)
ALT_ROW_FILL = PatternFill(fill_type="solid", fgColor=ROW_ALT_BG)

# ── Fonts ──────────────────────────────────────────────────────────────────────

HEADER_FONT = Font(bold=True, color=HEADER_FG, name="Calibri", size=11)
BOLD_FONT = Font(bold=True, name="Calibri", size=10)
NORMAL_FONT = Font(name="Calibri", size=10)

# ── Thin border ────────────────────────────────────────────────────────────────

_THIN = Side(style="thin", color="D9D9D9")
THIN_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

# ── Number formats ─────────────────────────────────────────────────────────────

FMT_BRL = '#,##0.0 "MM"'   # R$ millions (e.g. 1,234.5 MM)
FMT_PCT = '0.0"%"'          # percentage  (e.g. 38.7%)
FMT_MULTIPLE = '0.00"x"'    # leverage    (e.g. 1.99x)
FMT_INT = "#,##0"           # integer     (e.g. 915,400)
FMT_DATE = "YYYY-MM-DD"

# ── Column width helpers ───────────────────────────────────────────────────────

MIN_COL_WIDTH = 12
MAX_COL_WIDTH = 40


def auto_fit_columns(worksheet: Worksheet) -> None:
    """Adjust column widths to fit content (capped at :data:`MAX_COL_WIDTH`).

    Parameters
    ----------
    worksheet:
        An :class:`openpyxl.worksheet.worksheet.Worksheet` instance.
    """
    for col_cells in worksheet.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            try:
                cell_len = len(str(cell.value or ""))
                if cell_len > max_len:
                    max_len = cell_len
            except Exception:
                pass
        adjusted = min(max(max_len + 2, MIN_COL_WIDTH), MAX_COL_WIDTH)
        worksheet.column_dimensions[col_letter].width = adjusted


def style_header_cell(cell: "openpyxl.cell.cell.Cell", bold: bool = True) -> None:
    """Apply header styling to a single cell.

    Parameters
    ----------
    cell:
        An :class:`openpyxl.cell.cell.Cell` instance.
    bold:
        Whether to use the bold header font (default: ``True``).
    """
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT if bold else Font(bold=False, color=HEADER_FG, name="Calibri", size=11)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = THIN_BORDER


def style_data_cell(cell: "openpyxl.cell.cell.Cell", row_idx: int, number_format: str | None = None) -> None:
    """Apply alternating row styling to a data cell.

    Parameters
    ----------
    cell:
        An :class:`openpyxl.cell.cell.Cell` instance.
    row_idx:
        1-based row index (used for alternating row colour).
    number_format:
        Optional Excel number format string to apply.
    """
    if row_idx % 2 == 0:
        cell.fill = ALT_ROW_FILL
    cell.font = NORMAL_FONT
    cell.alignment = Alignment(horizontal="right", vertical="center")
    cell.border = THIN_BORDER
    if number_format:
        cell.number_format = number_format


def style_variation_cell(cell: "openpyxl.cell.cell.Cell", value: float | None, row_idx: int) -> None:
    """Apply coloured font to a variation cell (green = positive, red = negative).

    Parameters
    ----------
    cell:
        An :class:`openpyxl.cell.cell.Cell` instance.
    value:
        The numeric variation value (percentage points).
    row_idx:
        1-based row index for alternating background.
    """
    style_data_cell(cell, row_idx, number_format=FMT_PCT)
    if value is None:
        cell.font = Font(color=VAR_NEUTRAL_FG, name="Calibri", size=10)
    elif value > 0:
        cell.font = Font(color=VAR_POSITIVE_FG, name="Calibri", size=10)
    elif value < 0:
        cell.font = Font(color=VAR_NEGATIVE_FG, name="Calibri", size=10)
    else:
        cell.font = Font(color=VAR_NEUTRAL_FG, name="Calibri", size=10)
